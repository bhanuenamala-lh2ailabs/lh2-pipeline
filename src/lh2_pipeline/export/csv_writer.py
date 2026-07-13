"""Phase 5 export — the deliverable schema.

The original 14 columns are FIXED in order (do not reorder; "Contact Number"
appears twice intentionally). ``Email`` is appended at the END as column 15 — an
append-only change that preserves every existing column position.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Optional
from urllib.parse import quote_plus

from ..models import Company, Person

# Fixed schema — order matters; duplicate header "Contact Number" is intentional.
# New columns are APPENDED only (never inserted), so existing indices never shift.
COLUMNS = [
    "#",
    "Company",
    "Founder(s)",
    "Founder LinkedIn (verified)",
    "Contact Number",
    "SPOC 2 Linkedin",
    "Contact Number",
    "Incorp. Year",
    "HQ / India delivery",
    "Approx. Headcount",
    "Headcount source (approx.)",
    "Segment",
    "Status",
    "Notes",
    "Email",           # col 15 — primary founder email (SignalHire; sales verifies)
    "Size Bucket",     # col 16 — 1-100 / 100-500 / 500-1000 (approx, from headcount)
]


# --------------------------------------------------------------------------- #
def _founders_cell(people: list[Person]) -> str:
    parts = []
    for p in people[:2]:  # primary + SPOC2
        if not p.name or p.name == "(verify)":
            parts.append("(verify)")
        elif p.role:
            parts.append(f"{p.name} ({p.role})")
        else:
            parts.append(p.name)
    return "; ".join(parts) if parts else "(verify)"


def _li(person: Optional[Person]) -> str:
    # Show whatever LinkedIn we have for the founder (confirmed via namesake guard,
    # or Signalhire company-matched). Unverified ones are flagged in Notes.
    if person and person.linkedin_url:
        return person.linkedin_url
    return ""


# Friendly provenance labels for the Headcount-source column.
SOURCE_LABELS = {
    "goodfirms": "GoodFirms", "clutch": "Clutch", "techbehemoths": "TechBehemoths",
    "manifest": "The Manifest", "designrush": "DesignRush", "nasscom": "NASSCOM",
}


def _headcount_source(co: Company) -> str:
    """Where the headcount came from (the directory), not the band again."""
    labels = []
    for s in co.sources_json or []:
        key = str(s.get("source", "")).lower()
        lbl = SOURCE_LABELS.get(key, key.title() if key else "")
        if lbl and lbl not in labels:
            labels.append(lbl)
    base = "; ".join(labels)
    # If the listing carried a distinct raw size string, append it for context.
    raw = (co.size_source or "").strip()
    if raw and raw.replace(" ", "") != (co.size_band or "").replace(" ", ""):
        base = f"{base} ({raw})" if base else raw
    return base


def _phone(person: Optional[Person]) -> str:
    return (person.phone or "") if person else ""


def _email(person: Optional[Person]) -> str:
    return (person.email or "") if person else ""


def _strip_sources(note: Optional[str]) -> str:
    """Drop the machine-readable 'sources: a, b' segment, keep human caveats."""
    if not note:
        return ""
    segs = [s.strip() for s in note.split(";")]
    keep = [s for s in segs if s and not s.lower().startswith("sources:")]
    return "; ".join(keep)


def _notes_cell(co: Company, people: list[Person]) -> str:
    bits: list[str] = []
    if co.gate_reason:
        bits.append(co.gate_reason)
    for p in people[:2]:
        if not p.name or p.name == "(verify)":
            continue
        # founder provenance caveat: name only from Signalhire title-search → verify
        import re as _re
        m = _re.search(r"sources:\s*([a-z_,\s]+)", p.notes or "", _re.IGNORECASE)
        srcs = {s.strip().lower() for s in (m.group(1).split(",") if m else []) if s.strip()}
        if srcs and srcs <= {"signalhire"}:
            bits.append(f"founder '{p.name}' via Signalhire - verify")
        human = _strip_sources(p.notes)
        if human:
            bits.append(human)
    # dedupe preserving order
    seen, out = set(), []
    for b in bits:
        if b not in seen:
            seen.add(b)
            out.append(b)
    return " | ".join(out)


def _resolver_url(co: Company, resolver: str) -> str:
    if co.website:
        w = co.website
        return w if "://" in w else "https://" + w
    q = quote_plus(f"{co.company_name} {co.city or ''} software company".strip())
    return resolver.format(query=q)


def _company_cell(co: Company, hyperlinked: bool, resolver: str) -> str:
    if not hyperlinked:
        return co.company_name
    url = _resolver_url(co, resolver)
    safe_name = co.company_name.replace('"', '""')
    return f'=HYPERLINK("{url}","{safe_name}")'


def _row_for(co: Company, people: list[Person], idx: int, hyperlinked: bool, resolver: str) -> list:
    primary = people[0] if people else None
    spoc2 = people[1] if len(people) > 1 else None
    return [
        idx,
        _company_cell(co, hyperlinked, resolver),
        _founders_cell(people),
        _li(primary),
        _phone(primary),
        _li(spoc2),
        _phone(spoc2),
        co.founded_year or "",
        co.hq_country or "",
        co.size_band or "",
        _headcount_source(co),
        co.segment or "",
        co.status or "",
        _notes_cell(co, people),
        _email(primary),
        co.size_bucket or "",
    ]


def build_rows(store, hyperlinked: bool, resolver: str) -> list[list]:  # noqa: ANN001
    rows: list[list] = []
    idx = 0
    for co in store.iter_companies(gate_pass=True):
        people = store.people_for(co.domain)  # primary first
        idx += 1
        rows.append(_row_for(co, people, idx, hyperlinked, resolver))
    return rows


def _is_full(people: list[Person]) -> bool:
    """Four-field: primary founder has a real name AND LinkedIn AND phone AND email
    (SPOC2 optional). No verification gate — presence is enough; sales verifies."""
    if not people:
        return False
    p = people[0]
    return bool(p.name and p.name != "(verify)" and p.linkedin_url and p.phone and p.email)


def write_append_csv(
    store,                       # noqa: ANN001
    path: Path,
    template_header: list[str],
    start_index: int = 1,
    enriched_only: bool = True,
    require_founder: bool = False,
    require_full: bool = False,
    exclude_domains: Optional[set] = None,
    limit: Optional[int] = None,
    hyperlinked: bool = False,
    resolver: str = "",
) -> int:
    """Write a CSV matching ``template_header`` exactly so it appends onto an
    existing sheet. Numbering continues from ``start_index``.
      - enriched_only: only firms with people rows (default).
      - require_founder: only firms with a real (non-(verify)) founder.
      - require_full: only firms with primary founder name + LinkedIn + phone.
      - exclude_domains: skip these domains (e.g. already-delivered firms).
      - limit: stop after this many rows.
    """
    ncols = len(template_header)
    excl = exclude_domains or set()
    rows: list[list] = []
    idx = start_index - 1
    for co in store.iter_companies(gate_pass=True):
        if co.domain in excl:
            continue
        people = store.people_for(co.domain)
        if enriched_only and not people:
            continue
        if require_full and not _is_full(people):
            continue
        if require_founder and not any(p.name and p.name != "(verify)" for p in people):
            continue
        idx += 1
        row = _row_for(co, people, idx, hyperlinked, resolver)  # 14 cols
        if len(row) < ncols:                                    # pad to template width
            row = row + [""] * (ncols - len(row))
        elif len(row) > ncols:
            row = row[:ncols]
        rows.append(row)
        if limit and len(rows) >= limit:
            break

    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(template_header)
        w.writerows(rows)
    return len(rows)


def write_csv(store, path: Path, hyperlinked: bool, resolver: str) -> int:  # noqa: ANN001
    rows = build_rows(store, hyperlinked, resolver)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(COLUMNS)
        w.writerows(rows)
    return len(rows)


# --------------------------------------------------------------------------- #
# LinkedIn human-review sheet
# --------------------------------------------------------------------------- #
REVIEW_COLUMNS = ["Company", "Domain", "Founder name", "City", "LinkedIn"]


def write_linkedin_review(store, path: Path) -> int:  # noqa: ANN001
    """Rows missing a confirmed LinkedIn URL, formatted for Sales-Navigator filling."""
    path.parent.mkdir(parents=True, exist_ok=True)
    n = 0
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(REVIEW_COLUMNS)
        for co in store.iter_companies(gate_pass=True):
            for p in store.people_for(co.domain):
                if p.is_primary and not p.linkedin_confirmed and p.name and p.name != "(verify)":
                    w.writerow([co.company_name, co.domain, p.name, co.city or "", ""])
                    n += 1
    return n


# --------------------------------------------------------------------------- #
# Optional XLSX in LH2 house style
# --------------------------------------------------------------------------- #
def write_xlsx(store, path: Path, hyperlinked: bool, resolver: str, style) -> int:  # noqa: ANN001
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as e:  # pragma: no cover
        raise RuntimeError('openpyxl not installed. Install with: pip install -e ".[xlsx]"') from e

    rows = build_rows(store, hyperlinked, resolver)
    wb = Workbook()
    ws = wb.active
    ws.title = "Targets"

    header_fill = PatternFill("solid", fgColor=style.header_fill)
    band_fill = PatternFill("solid", fgColor=style.band_fill)
    header_font = Font(name=style.font, size=style.font_size, bold=True, color="FFFFFF")
    body_font = Font(name=style.font, size=style.font_size)

    ws.append(COLUMNS)
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for i, row in enumerate(rows, start=2):
        ws.append(row)
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=i, column=c)
            cell.font = body_font
            if i % 2 == 0:
                cell.fill = band_fill

    # freeze header (freeze panes at A2 keeps row 1 visible), auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return len(rows)
