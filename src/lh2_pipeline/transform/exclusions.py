"""Net-new exclusion: harvest already-known / previously-mined firms so the
build phase never re-emits them.

Two independent signals, either of which excludes a firm at the gate:
  * **domain** — the canonical registered domain (the dedupe key). Strongest.
  * **name**   — distinctive-core company-name match (see gates.matches_known_firm).

Sources are config-driven (``gates.exclude_name_files`` / ``exclude_domain_files``
plus the legacy singular ``blocklist_known_file``). Formats handled per file:
  * ``.txt``  — one domain per line (``data/delivered_domains.txt``).
  * ``.csv``  — a ``Company`` / ``Company Name`` column for names; a
    ``Domain`` / ``Website`` / ``URL`` column for domains; and, crucially, a
    Google-Sheets ``=HYPERLINK("https://firm.com/","Firm Name")`` formula in the
    Company column (how ``append_ready.csv`` / ``targets_hyperlinked.csv`` store
    the delivered firms) — both the display name and the domain are extracted.

Never guesses: a cell that yields no registrable domain contributes only its name.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field

from ..logging_setup import get_logger
from .canonicalize import canonical_domain

log = get_logger("lh2.exclusions")

# =HYPERLINK("url","display")  — quotes already un-doubled by csv.reader.
_HYPERLINK_RE = re.compile(r'=HYPERLINK\(\s*"([^"]+)"\s*,\s*"([^"]*)"\s*\)', re.IGNORECASE)

_NAME_HEADERS = {"company", "company name", "firm", "name"}
_DOMAIN_HEADERS = {"domain", "domain name", "website", "url", "site", "web"}


@dataclass
class Exclusions:
    names: list[str] = field(default_factory=list)      # for fuzzy core-name match
    domains: set[str] = field(default_factory=set)      # canonical domains

    def merge(self, other: "Exclusions") -> None:
        self.names.extend(other.names)
        self.domains |= other.domains


def _unwrap_hyperlink(cell: str) -> tuple[str | None, str | None]:
    """Return (display_name, url) from a =HYPERLINK() cell, else (cell, None)."""
    m = _HYPERLINK_RE.search(cell)
    if m:
        return (m.group(2).strip() or None, m.group(1).strip() or None)
    return (cell.strip() or None, None)


def _col_index(header: list[str], candidates: set[str]) -> int | None:
    for i, h in enumerate(header):
        if h.strip().lower() in candidates:
            return i
    return None


def load_domain_file(path) -> set[str]:  # noqa: ANN001
    """One-domain-per-line .txt -> canonical domain set."""
    out: set[str] = set()
    with open(path, "r", encoding="utf-8-sig") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            d = canonical_domain(line)
            if d:
                out.add(d)
    return out


def _harvest_row(ex: "Exclusions", row, name_i, domain_i) -> None:  # noqa: ANN001
    """Pull a name (+ any embedded/domain-like domain) and a domain-column value
    from one row into ``ex``."""
    if name_i is not None and name_i < len(row) and row[name_i] and str(row[name_i]).strip():
        display, url = _unwrap_hyperlink(str(row[name_i]))
        if display:
            ex.names.append(display)
            # a domain-like company name (e.g. "Supersei.ai") is also a domain
            if " " not in display:
                d = canonical_domain(display)
                if d:
                    ex.domains.add(d)
        if url:
            d = canonical_domain(url)
            if d:
                ex.domains.add(d)
    if domain_i is not None and domain_i < len(row) and row[domain_i] and str(row[domain_i]).strip():
        d = canonical_domain(str(row[domain_i]))
        if d:
            ex.domains.add(d)


def load_name_xlsx(path) -> Exclusions:  # noqa: ANN001
    """Harvest names + domains from every sheet of an .xlsx workbook. Each sheet's
    header is scanned for a Company / Domain column independently."""
    ex = Exclusions()
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.info("exclude_xlsx_needs_openpyxl", path=str(path))
        return ex
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            header = next(it, None)
            if not header:
                continue
            hdr = [str(c).strip().lower() if c is not None else "" for c in header]
            name_i = next((i for i, h in enumerate(hdr) if h in _NAME_HEADERS), None)
            domain_i = next((i for i, h in enumerate(hdr) if h in _DOMAIN_HEADERS), None)
            if name_i is None and domain_i is None:
                continue
            for row in it:
                _harvest_row(ex, row, name_i, domain_i)
    finally:
        wb.close()
    return ex


def load_name_file(path) -> Exclusions:  # noqa: ANN001
    """Dispatch a name/known file to the CSV or XLSX harvester by extension."""
    if path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        return load_name_xlsx(path)
    return load_name_csv(path)


def load_name_csv(path) -> Exclusions:  # noqa: ANN001
    """Harvest names + domains from a delivery/known CSV."""
    ex = Exclusions()
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        header = next(reader, None)
        if not header:
            return ex
        name_i = _col_index(header, _NAME_HEADERS)
        if name_i is None:
            name_i = 1 if len(header) > 1 else 0        # legacy default: 2nd col
        domain_i = _col_index(header, _DOMAIN_HEADERS)
        for row in reader:
            if name_i < len(row) and row[name_i].strip():
                display, url = _unwrap_hyperlink(row[name_i])
                if display:
                    ex.names.append(display)
                if url:
                    d = canonical_domain(url)
                    if d:
                        ex.domains.add(d)
            if domain_i is not None and domain_i < len(row) and row[domain_i].strip():
                d = canonical_domain(row[domain_i])
                if d:
                    ex.domains.add(d)
    return ex


def load_exclusions(cfg) -> Exclusions:  # noqa: ANN001
    """Aggregate every configured known/mined source into one Exclusions set.

    Includes the legacy singular ``blocklist_known_file`` and the config
    ``blocklist_known_names`` so this is the single source of truth for run_build.
    """
    gates = cfg.gates
    ex = Exclusions()

    # Guard against the self-referential footgun: the export writer regenerates
    # files under exports_dir every run, so using one as an exclusion source would
    # exclude the entire candidate list. Skip anything inside exports_dir.
    exports_dir = None
    try:
        exports_dir = cfg.exports_dir.resolve()
    except Exception:
        exports_dir = None

    def _is_export_output(path) -> bool:  # noqa: ANN001
        if exports_dir is None:
            return False
        try:
            path.resolve().relative_to(exports_dir)
            return True
        except Exception:
            return False

    # config-level literals
    ex.names.extend(n for n in (gates.blocklist_known_names or []) if n and n.strip())
    for d in (gates.blocklist_known_domains or []):
        cd = canonical_domain(d)
        if cd:
            ex.domains.add(cd)

    # name/known CSVs (legacy singular + new plural list), de-duplicated by path
    name_files: list[str] = []
    if getattr(gates, "blocklist_known_file", None):
        name_files.append(gates.blocklist_known_file)
    name_files.extend(getattr(gates, "exclude_name_files", []) or [])
    seen_paths: set[str] = set()
    for rel in name_files:
        path = cfg.abspath(rel)
        key = str(path)
        if key in seen_paths:
            continue
        seen_paths.add(key)
        if _is_export_output(path):
            log.info("exclude_file_skipped_export_output", path=key)
            continue
        if not path.exists():
            log.info("exclude_file_missing", path=key)
            continue
        try:
            ex.merge(load_name_file(path))
        except Exception as e:  # never let one malformed file break the build
            log.info("exclude_file_error", path=key, err=str(e))

    # domain-only files (.txt one-per-line, or CSV with a domain column)
    for rel in (getattr(gates, "exclude_domain_files", []) or []):
        path = cfg.abspath(rel)
        if _is_export_output(path):
            log.info("exclude_file_skipped_export_output", path=str(path))
            continue
        if not path.exists():
            log.info("exclude_file_missing", path=str(path))
            continue
        try:
            if path.suffix.lower() == ".csv":
                ex.merge(load_name_csv(path))
            else:
                ex.domains |= load_domain_file(path)
        except Exception as e:
            log.info("exclude_file_error", path=str(path), err=str(e))

    # de-dup names while preserving order
    seen_n: set[str] = set()
    deduped: list[str] = []
    for n in ex.names:
        k = n.strip().lower()
        if k and k not in seen_n:
            seen_n.add(k)
            deduped.append(n)
    ex.names = deduped

    log.info("exclusions_loaded", names=len(ex.names), domains=len(ex.domains))
    return ex
